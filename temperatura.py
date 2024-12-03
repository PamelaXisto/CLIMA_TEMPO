from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from tkinter import *
from datetime import datetime
import os

#biblioteca
from openpyxl import load_workbook, Workbook

def buscar_previsao():

    navegador = webdriver.Chrome()

    #Acessa o google para fazer a pesquisa
    navegador.get('https://www.google.com')

    #É digitado a informação que deseja
    navegador.find_element(By.XPATH, '//*[@id="APjFqb"]').send_keys('temperatura de são paulo')

    navegador.find_element(By.XPATH, '/html/body/div[1]/div[3]/form/div[1]/div[1]/div[3]/center/input[1]').send_keys(Keys.ENTER)

    #Temperatura e Umidade em tempo real
    temperatura = navegador.find_element(By.XPATH, '//*[@id="wob_tm"]').text
    umidade = navegador.find_element(By.XPATH, '//*[@id="wob_wc"]/div[1]/div[2]/div[2]').text
    umidade = umidade.replace("Umidade: ", "").replace("%", "")

    print(f'TEMPERATURA ATUAL DE SÃO PAULO: {temperatura}°C')
    print(f'UMIDADE: {umidade}%')
    
    #Momento em que fecha o navegador
    navegador.quit()

    arquivo = 'captacao.xlsx'
    if not os.path.exists(arquivo):

        # Cria um novo arquivo Excel se não existir
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Clima"

        # Adiciona os cabeçalhos em cada coluna
        sheet.append(["DATA", "HORA", "TEMPERATURA (°C)", "UMIDADE (%)"])
        wb.save(arquivo)

    # Carrega o arquivo Excel
    arquivo = load_workbook('captacao.xlsx')
    sheet = arquivo.active

    # Adiciona atualizações dos dados de temperatura na próxima linha/entrada de dados
    data = datetime.now().strftime("%Y-%m-%d")
    hora = datetime.now().strftime("%H:%M:%S")
    sheet.append([data, hora, temperatura, umidade])

    # Salva o arquivo com as novas informações
    arquivo.save('captacao.xlsx')
    print(f"Planilha atualizada com sucesso!!!")    

root = Tk()
#Título da minha janela do Tkinter
root.title("Clima tempo de São Paulo")

#Definição da altura / largura da janela Tkinter
root.geometry("350x100")

#Texto acima do botão
myLabel = Label(root, text="Atualizar a previsão do tempo!")
myLabel.pack(pady=10)

#Botão para solicitar o clima tempo
myButton = Button(root, text="Buscar previsão", command=buscar_previsao)
myButton.pack(pady=5)

root.mainloop()

